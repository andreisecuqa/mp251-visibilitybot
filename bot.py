"""
Telegram Bot - Managementul Vizibilității Exam Generator
Folosește Gemini API pentru generarea conținutului + openpyxl pentru export Excel
"""

import os
import logging
import asyncio
import copy
import shutil
from pathlib import Path
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, ConversationHandler, ContextTypes, filters
)
import requests
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

# ─── CONFIG ────────────────────────────────────────────────────────────────────
TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "YOUR_BOT_TOKEN")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "YOUR_GEMINI_KEY")
TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"
OUTPUT_DIR = Path(__file__).parent / "outputs"
OUTPUT_DIR.mkdir(exist_ok=True)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# ─── CONVERSATION STATES ───────────────────────────────────────────────────────
(
    ASK_NAME, ASK_FIRSTNAME, ASK_ORGANIZATION, ASK_DOMAIN,
    ASK_ORG_TYPE, ASK_MARKET, ASK_COMPETITORS, ASK_PLATFORMS,
    ASK_GRUPA, CONFIRM, GENERATING
) = range(11)

# ─── GEMINI SETUP ──────────────────────────────────────────────────────────────
GEMINI_BASE = "https://generativelanguage.googleapis.com/v1/models"

def gemini_generate(prompt: str) -> str:
    """Apel REST direct la Gemini API v1 - fara SDK, fara v1beta."""
    # Lista modele disponibile via v1
    r = requests.get(f"{GEMINI_BASE}?key={GEMINI_API_KEY}", timeout=10)
    r.raise_for_status()
    models = r.json().get("models", [])
    
    # Filtrare modele care suporta generateContent
    available = [
        m["name"].replace("models/", "")
        for m in models
        if "generateContent" in m.get("supportedGenerationMethods", [])
    ]
    logger.info(f"Modele disponibile v1: {available}")

    # Ordinea de preferinta - modele stabile > modele noi (mai putin stabile)
    preferred = ["gemini-2.0-flash-001", "gemini-2.0-flash", "gemini-2.0-flash-lite", "gemini-2.5-flash", "gemini-2.5-flash-lite", "gemini-2.5-pro"]
    ordered = []
    for p in preferred:
        for a in available:
            if a == p and a not in ordered:
                ordered.append(a)
    # adauga restul care nu sunt in lista de preferinte
    for a in available:
        if a not in ordered:
            ordered.append(a)

    if not ordered:
        raise Exception("Niciun model disponibil.")

    payload = {"contents": [{"parts": [{"text": prompt}]}]}
    last_error = None
    for chosen in ordered:
        url = f"{GEMINI_BASE}/{chosen}:generateContent?key={GEMINI_API_KEY}"
        logger.info(f"Incerc modelul: {chosen}")
        try:
            resp = requests.post(url, json=payload, timeout=120)
            if resp.status_code in (429, 500, 503):
                logger.warning(f"Model {chosen} a returnat {resp.status_code}, incerc urmatorul")
                last_error = f"{resp.status_code} {resp.text[:100]}"
                continue
            resp.raise_for_status()
            logger.info(f"Succes cu modelul: {chosen}")
            return resp.json()["candidates"][0]["content"]["parts"][0]["text"]
        except Exception as e:
            logger.warning(f"Model {chosen} a esuat: {e}")
            last_error = e
            continue

    raise Exception(f"Toate modelele au esuat. Ultima eroare: {last_error}")

# ─── SYSTEM PROMPT - antidetecție AI ──────────────────────────────────────────
SYSTEM_PERSONA = """
Ești un student masterand în Managementul Proiectelor care completează un proiect de examen.
Scrie EXCLUSIV în română. Stilul tău:
- Propoziții medii (12-20 cuvinte), uneori mai scurte
- Ocazional o frază mai lungă, dar nu exagerat de academică
- Folosești termenii din curs natural, nu forțat
- Ești specific și concret, nu generic
- Dai exemple reale, cu cifre plauzibile
- Uneori o ușoară nesiguranță ("consider că", "se poate observa că", "după cum reiese")
- NU folosești bullet points cu simboluri speciale — folosești doar text structurat cu numerotare simplă
- NU ești perfect — câte o virgulă în plus, o frază mai puțin elegantă
- NU scrii ca un raport corporatist — ești student, nu consultant McKinsey
- Folosește "organizația" nu "brandul", "postare" nu "content", "urmăritori" nu "followeri" (decât pe LinkedIn)
- Evită: "în concluzie", "este esențial să", "în lumina celor prezentate", "cu titlu de exemplu"
- Preferă: "din ce am analizat", "se vede clar că", "ceea ce funcționează bine este", "un lucru care lipsește"
"""

# ─── EXCEL WRITER ──────────────────────────────────────────────────────────────
def fill_excel(student_data: dict, generated: dict, output_path: Path):
    """Completează template-ul Excel cu datele generate."""
    shutil.copy(TEMPLATE_PATH, output_path)
    wb = load_workbook(output_path)

    sheet_map = {
        "GHID": fill_ghid,
        "BRIEF": fill_brief,
        "SOCIAL MEDIA ANALYSIS": fill_social_media,
        "ANALIZA VIZIBILITĂȚII": fill_analiza_vizib,
        "CONTENT CU IMPACT RIDICAT": fill_content_impact,
        "BANNERE ȘI VIZUALURI": fill_bannere,
        "CLIENT TINTA": fill_client,
        "STRATEGIE DE VIZIBILITATE": fill_strategie,
        "COPYWRITING": fill_copywriting,
        "LINKEDIN": fill_linkedin,
        "MATERIALE CANVA": fill_canva,
        "CONTENT PLAN": fill_content_plan,
    }

    for sheet_name, filler_fn in sheet_map.items():
        if sheet_name in wb.sheetnames:
            try:
                filler_fn(wb[sheet_name], student_data, generated)
            except Exception as e:
                logger.error(f"Eroare la sheet {sheet_name}: {e}")

    wb.save(output_path)
    return output_path


def set_cell(ws, row, col, value, bold=False, wrap=True, size=10):
    """Helper pentru a seta o celulă cu formatare."""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=bold, size=size, name="Calibri")
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    return cell


def fill_ghid(ws, sd, gen):
    # Datele studentului în GHID
    pairs = [
        (6, 3, sd["last_name"]),
        (7, 3, sd["first_name"]),
        (9, 3, sd["grupa"]),
        (10, 3, f"{sd['org_name']}"),
        (11, 3, sd["domain"]),
        (12, 3, datetime.now().strftime("%d/%m/%Y")),
    ]
    for row, col, val in pairs:
        set_cell(ws, row, col, val)


def fill_brief(ws, sd, gen):
    data = gen.get("brief", {})
    mappings = [
        (6, 3, sd["org_name"]),
        (7, 3, sd["domain"]),
        (8, 3, sd["org_type"]),
        (9, 3, data.get("an_infiintare", "2015")),
        (10, 3, data.get("misiune", "")),
        (11, 3, data.get("viziune", "")),
        (12, 3, data.get("valori", "")),
        (13, 3, data.get("puv", "")),
        (14, 3, data.get("arie_geografica", "")),
        (15, 3, data.get("website", f"https://www.{sd['org_name'].lower().replace(' ', '')}.md")),
        (18, 3, data.get("segment_piata", "")),
        (19, 3, data.get("pozitionare_concurenti", "")),
        (20, 3, sd.get("competitors", "")),
        (21, 3, data.get("swot", "")),
        (26, 3, data.get("content_informativ", "")),
        (27, 3, data.get("content_educational", "")),
        (28, 3, data.get("content_promotional", "")),
        (29, 3, data.get("content_engagement", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)


def fill_social_media(ws, sd, gen):
    data = gen.get("social_media", {})
    platforms = ["Facebook", "Instagram", "LinkedIn", "TikTok", "YouTube"]
    rows = [15, 16, 17, 18, 19]
    for i, (plat, row) in enumerate(zip(platforms, rows)):
        key = plat.lower()
        set_cell(ws, row, 3, data.get(f"{key}_content", ""))
        set_cell(ws, row, 4, data.get(f"{key}_frecventa", ""))
        set_cell(ws, row, 5, data.get(f"{key}_engagement", ""))
        set_cell(ws, row, 6, data.get(f"{key}_observatii", ""))

    set_cell(ws, 25, 2, data.get("top3_facebook", ""))
    set_cell(ws, 27, 2, data.get("top3_instagram", ""))
    set_cell(ws, 29, 2, data.get("top3_linkedin", ""))


def fill_analiza_vizib(ws, sd, gen):
    data = gen.get("analiza_vizib", {})
    mappings = [
        (6, 3, data.get("cum_construita", "")),
        (7, 3, data.get("canale_proiecte", "")),
        (8, 3, data.get("omnichannel", "")),
        (9, 3, data.get("seo", "")),
        (10, 3, data.get("paid_ads", "")),
        (11, 3, data.get("mkt_situational", "")),
        (14, 3, data.get("consistenta_mesaje", "")),
        (15, 3, data.get("coerenta_vizuala", "")),
        (16, 3, data.get("branding_proiecte", "")),
        (17, 3, data.get("recunoastere_vizuala", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)

    puncte_forte = data.get("puncte_forte", [])
    lacune = data.get("lacune", [])
    for i in range(6):
        r = 21 + i
        set_cell(ws, r, 2, puncte_forte[i] if i < len(puncte_forte) else "")
        set_cell(ws, r, 3, lacune[i] if i < len(lacune) else "")


def fill_content_impact(ws, sd, gen):
    data = gen.get("content_impact", [])
    start_rows = [12, 25, 38]
    for idx, (postare, start) in enumerate(zip(data[:3], start_rows)):
        set_cell(ws, start + 1, 3, postare.get("link", f"https://facebook.com/{sd['org_name'].replace(' ', '')}"))
        set_cell(ws, start + 2, 3, postare.get("platforma", "Facebook"))
        set_cell(ws, start + 3, 3, postare.get("tip_continut", "Imagine"))
        set_cell(ws, start + 4, 3, postare.get("data", "01/03/2025"))
        set_cell(ws, start + 5, 3, postare.get("metrici", ""))
        set_cell(ws, start + 6, 3, postare.get("factor_vizual", ""))
        set_cell(ws, start + 7, 3, postare.get("factor_mesaj", ""))
        set_cell(ws, start + 8, 3, postare.get("factor_timing", ""))
        set_cell(ws, start + 9, 3, postare.get("factor_distributie", ""))
        set_cell(ws, start + 10, 3, postare.get("elemente_replicabile", ""))


def fill_bannere(ws, sd, gen):
    data = gen.get("bannere", {})
    mappings = [
        (6, 3, data.get("culori_principale", "")),
        (7, 3, data.get("culori_secundare", "")),
        (8, 3, data.get("consecventa_culori", "")),
        (9, 3, data.get("psihologia_culorilor", "")),
        (10, 3, data.get("contrast", "")),
        (11, 3, data.get("fonturi_titluri", "")),
        (12, 3, data.get("fonturi_body", "")),
        (13, 3, data.get("fonturi_consecvente", "")),
        (16, 3, data.get("structura_bannere", "")),
        (17, 3, data.get("structura_social", "")),
        (18, 3, data.get("spatiu_negativ", "")),
        (19, 3, data.get("ierarhie_vizuala", "")),
        (20, 3, data.get("tipografii", "")),
        (23, 3, data.get("logo_prezent", "")),
        (24, 3, data.get("logo_corect", "")),
        (25, 3, data.get("identitate_unitara", "")),
        (26, 3, data.get("elemente_recurente", "")),
        (27, 3, data.get("personalitate_brand", "")),
        (28, 3, data.get("calitate_vizuala", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)


def fill_client(ws, sd, gen):
    data = gen.get("client_tinta", {})
    mappings = [
        (9, 3, data.get("varsta", "")),
        (10, 3, data.get("gen", "")),
        (11, 3, data.get("locatie", "")),
        (12, 3, data.get("educatie", "")),
        (13, 3, data.get("venit", "")),
        (14, 3, data.get("stare_civila", "")),
        (18, 3, data.get("ocupatie", "")),
        (19, 3, data.get("sector", "")),
        (20, 3, data.get("nivel_ierarhic", "")),
        (21, 3, data.get("tip_org", "")),
        (25, 3, data.get("probleme_principale", "")),
        (26, 3, data.get("nevoi_nerezolvate", "")),
        (27, 3, data.get("obiectii", "")),
        (28, 3, data.get("dream_outcome", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)


def fill_strategie(ws, sd, gen):
    data = gen.get("strategie", {})
    mappings = [
        (7, 3, data.get("obiectiv1", "")),
        (8, 3, data.get("obiectiv2", "")),
        (9, 3, data.get("obiectiv3", "")),
        (10, 3, data.get("obiectiv4", "")),
        (14, 3, data.get("platforma1", "")),
        (15, 3, data.get("platforma2", "")),
        (16, 3, data.get("platforma3", "")),
        (17, 3, data.get("platforma_renuntat", "")),
        (21, 3, data.get("content_educational", "")),
        (22, 3, data.get("content_engagement", "")),
        (23, 3, data.get("content_storytelling", "")),
        (24, 3, data.get("content_promotional", "")),
        (25, 3, data.get("content_entertainment", "")),
        (28, 3, data.get("directia1", "")),
        (29, 3, data.get("directia2", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)


def fill_copywriting(ws, sd, gen):
    data = gen.get("copywriting", [])
    start_row = 25
    for i, text in enumerate(data[:6]):
        row = start_row + i
        set_cell(ws, row, 2, text.get("platforma", ""))
        set_cell(ws, row, 3, text.get("hook", ""))
        set_cell(ws, row, 4, text.get("mesaj", ""))
        set_cell(ws, row, 5, text.get("cta", ""))


def fill_linkedin(ws, sd, gen):
    data = gen.get("linkedin", {})
    mappings = [
        (11, 3, sd["org_name"]),
        (12, 3, data.get("headline", "")),
        (13, 3, data.get("banner_descriere", "")),
        (14, 3, data.get("buton_cta", "")),
        (18, 3, data.get("despre", "")),
        (22, 3, data.get("specialitati", "")),
        (23, 3, data.get("website", "")),
        (24, 3, data.get("industria", "")),
        (25, 3, data.get("marime", "")),
        (26, 3, data.get("tip_companie", "")),
        (27, 3, data.get("an_infiintare", "")),
        (28, 3, data.get("locatie", "")),
    ]
    for row, col, val in mappings:
        set_cell(ws, row, col, val)


def fill_canva(ws, sd, gen):
    data = gen.get("canva", {})
    pairs = [
        (16, "informativ"),
        (19, "promotional"),
        (22, "educational"),
        (25, "engagement"),
    ]
    for row, key in pairs:
        item = data.get(key, {})
        set_cell(ws, row, 2, item.get("descriere", ""))
        set_cell(ws, row, 3, item.get("copy", ""))
        set_cell(ws, row, 4, f"[Screenshot material {key} inserat manual]")


def fill_content_plan(ws, sd, gen):
    data = gen.get("content_plan", {})
    set_cell(ws, 6, 4, data.get("obiectiv_principal", ""))
    set_cell(ws, 7, 4, data.get("tema_lunii", ""))
    set_cell(ws, 8, 4, data.get("frecventa", ""))
    set_cell(ws, 9, 4, data.get("mix_continut", ""))
    set_cell(ws, 10, 4, data.get("tonalitate", ""))

    # Calendarul săptămânal
    weeks = data.get("saptamani", [])
    for i, week in enumerate(weeks[:4]):
        row = 14 + i
        zile = ["luni", "marti", "miercuri", "joi", "vineri", "sambata", "duminica"]
        for j, zi in enumerate(zile):
            set_cell(ws, row, 2 + j, week.get(zi, ""))

    # Tabelul detaliat
    postari = data.get("postari_detaliate", [])
    for i, postare in enumerate(postari[:8]):
        row = 23 + i
        set_cell(ws, row, 2, postare.get("data", ""))
        set_cell(ws, row, 3, postare.get("platforma", ""))
        set_cell(ws, row, 4, postare.get("tip", ""))
        set_cell(ws, row, 5, postare.get("topic", ""))
        set_cell(ws, row, 6, postare.get("format", ""))
        set_cell(ws, row, 7, postare.get("copy_hook", ""))
        set_cell(ws, row, 8, postare.get("kpi", ""))


# ─── GEMINI GENERATOR ──────────────────────────────────────────────────────────
def generate_with_gemini(student_data: dict) -> dict:
    """Generează tot conținutul pentru examen cu un singur apel Gemini structurat."""

    org = student_data["org_name"]
    domain = student_data["domain"]
    org_type = student_data["org_type"]
    market = student_data["market"]
    competitors = student_data.get("competitors", "concurenții din piață")
    platforms = student_data.get("platforms", "Facebook, Instagram, LinkedIn")

    prompt = f"""
{SYSTEM_PERSONA}

Ești student masterand și completezi un proiect de examen la disciplina "Managementul Vizibilității Proiectelor".
Organizația pe care o analizezi: "{org}"
Domeniul: "{domain}"
Tipul organizației: "{org_type}"
Piața / contextul: "{market}"
Competitori menționați: "{competitors}"
Platformele active: "{platforms}"

Generează conținut COMPLET și PLAUZIBIL pentru fiecare secțiune. Inventează cifre realiste, exemple concrete, date specifice. Tot ce scrii trebuie să pară că a fost cercetat și scris de un student.

Răspunde STRICT în format JSON cu structura de mai jos. NU adăuga text în afara JSON-ului.

{{
  "brief": {{
    "an_infiintare": "...",
    "misiune": "...",
    "viziune": "...",
    "valori": "...",
    "puv": "...",
    "arie_geografica": "...",
    "website": "https://...",
    "segment_piata": "...",
    "pozitionare_concurenti": "...",
    "swot": "Puncte forte: ... | Puncte slabe: ... | Oportunități: ... | Amenințări: ...",
    "content_informativ": "...",
    "content_educational": "...",
    "content_promotional": "...",
    "content_engagement": "..."
  }},
  "social_media": {{
    "facebook_content": "...",
    "facebook_frecventa": "...",
    "facebook_engagement": "...",
    "facebook_observatii": "...",
    "instagram_content": "...",
    "instagram_frecventa": "...",
    "instagram_engagement": "...",
    "instagram_observatii": "...",
    "linkedin_content": "...",
    "linkedin_frecventa": "...",
    "linkedin_engagement": "...",
    "linkedin_observatii": "...",
    "tiktok_content": "...",
    "tiktok_frecventa": "...",
    "tiktok_engagement": "...",
    "tiktok_observatii": "...",
    "youtube_content": "...",
    "youtube_frecventa": "...",
    "youtube_engagement": "...",
    "youtube_observatii": "...",
    "top3_facebook": "1. [link + tip conținut + engagement + de ce a performat]\n2. ...\n3. ...",
    "top3_instagram": "1. ...\n2. ...\n3. ...",
    "top3_linkedin": "1. ...\n2. ...\n3. ..."
  }},
  "analiza_vizib": {{
    "cum_construita": "...",
    "canale_proiecte": "...",
    "omnichannel": "...",
    "seo": "...",
    "paid_ads": "...",
    "mkt_situational": "...",
    "consistenta_mesaje": "...",
    "coerenta_vizuala": "...",
    "branding_proiecte": "...",
    "recunoastere_vizuala": "...",
    "puncte_forte": ["...", "...", "...", "...", "...", "..."],
    "lacune": ["...", "...", "...", "...", "...", "..."]
  }},
  "content_impact": [
    {{
      "link": "https://...",
      "platforma": "...",
      "tip_continut": "...",
      "data": "ZZ/LL/AAAA",
      "metrici": "Like-uri: X | Share-uri: X | Comentarii: X | Reach: X | Engagement Rate: X%",
      "factor_vizual": "...",
      "factor_mesaj": "...",
      "factor_timing": "...",
      "factor_distributie": "...",
      "elemente_replicabile": "1. ...\n2. ...\n3. ..."
    }},
    {{ ... }},
    {{ ... }}
  ],
  "bannere": {{
    "culori_principale": "...",
    "culori_secundare": "...",
    "consecventa_culori": "...",
    "psihologia_culorilor": "...",
    "contrast": "...",
    "fonturi_titluri": "...",
    "fonturi_body": "...",
    "fonturi_consecvente": "...",
    "structura_bannere": "...",
    "structura_social": "...",
    "spatiu_negativ": "...",
    "ierarhie_vizuala": "...",
    "tipografii": "...",
    "logo_prezent": "...",
    "logo_corect": "...",
    "identitate_unitara": "...",
    "elemente_recurente": "...",
    "personalitate_brand": "...",
    "calitate_vizuala": "..."
  }},
  "client_tinta": {{
    "varsta": "...",
    "gen": "...",
    "locatie": "...",
    "educatie": "...",
    "venit": "...",
    "stare_civila": "...",
    "ocupatie": "...",
    "sector": "...",
    "nivel_ierarhic": "...",
    "tip_org": "...",
    "probleme_principale": "1. ...\n2. ...\n3. ...",
    "nevoi_nerezolvate": "...",
    "obiectii": "...",
    "dream_outcome": "..."
  }},
  "strategie": {{
    "obiectiv1": "...",
    "obiectiv2": "...",
    "obiectiv3": "...",
    "obiectiv4": "...",
    "platforma1": "...",
    "platforma2": "...",
    "platforma3": "...",
    "platforma_renuntat": "...",
    "content_educational": "40% - ...",
    "content_engagement": "20% - ...",
    "content_storytelling": "20% - ...",
    "content_promotional": "15% - ...",
    "content_entertainment": "5% - ...",
    "directia1": "...",
    "directia2": "..."
  }},
  "copywriting": [
    {{"platforma": "Facebook - Promoțional", "hook": "...", "mesaj": "...", "cta": "..."}},
    {{"platforma": "Instagram - Educațional", "hook": "...", "mesaj": "...", "cta": "..."}},
    {{"platforma": "LinkedIn - Profesional", "hook": "...", "mesaj": "...", "cta": "..."}},
    {{"platforma": "TikTok - Engagement", "hook": "...", "mesaj": "...", "cta": "..."}},
    {{"platforma": "Email - Newsletter", "hook": "...", "mesaj": "...", "cta": "..."}}
  ],
  "linkedin": {{
    "headline": "...",
    "banner_descriere": "...",
    "buton_cta": "...",
    "despre": "...",
    "specialitati": "...",
    "website": "https://...",
    "industria": "...",
    "marime": "...",
    "tip_companie": "...",
    "an_infiintare": "...",
    "locatie": "Chișinău, Republica Moldova"
  }},
  "canva": {{
    "informativ": {{"descriere": "...", "copy": "..."}},
    "promotional": {{"descriere": "...", "copy": "..."}},
    "educational": {{"descriere": "...", "copy": "..."}},
    "engagement": {{"descriere": "...", "copy": "..."}}
  }},
  "content_plan": {{
    "obiectiv_principal": "...",
    "tema_lunii": "...",
    "frecventa": "...",
    "mix_continut": "...",
    "tonalitate": "...",
    "saptamani": [
      {{"luni": "...", "marti": "...", "miercuri": "...", "joi": "...", "vineri": "...", "sambata": "...", "duminica": "Pauză"}},
      {{"luni": "...", "marti": "...", "miercuri": "...", "joi": "...", "vineri": "...", "sambata": "...", "duminica": "Pauză"}},
      {{"luni": "...", "marti": "...", "miercuri": "...", "joi": "...", "vineri": "...", "sambata": "...", "duminica": "Pauză"}},
      {{"luni": "...", "marti": "...", "miercuri": "...", "joi": "...", "vineri": "...", "sambata": "...", "duminica": "Pauză"}}
    ],
    "postari_detaliate": [
      {{"data": "01.05", "platforma": "Instagram", "tip": "Educational", "topic": "...", "format": "Carousel", "copy_hook": "...", "kpi": "..."}},
      {{"data": "03.05", "platforma": "LinkedIn", "tip": "Thought Leadership", "topic": "...", "format": "Articol + imagine", "copy_hook": "...", "kpi": "..."}},
      {{"data": "05.05", "platforma": "Facebook", "tip": "Promoțional", "topic": "...", "format": "Video 60 sec", "copy_hook": "...", "kpi": "..."}},
      {{"data": "08.05", "platforma": "Instagram", "tip": "Storytelling", "topic": "...", "format": "Reels 45 sec", "copy_hook": "...", "kpi": "..."}},
      {{"data": "10.05", "platforma": "TikTok", "tip": "Engagement", "topic": "...", "format": "Video scurt", "copy_hook": "...", "kpi": "..."}},
      {{"data": "13.05", "platforma": "Facebook", "tip": "Educational", "topic": "...", "format": "Infografic", "copy_hook": "...", "kpi": "..."}},
      {{"data": "15.05", "platforma": "LinkedIn", "tip": "Promoțional", "topic": "...", "format": "Post text + imagine", "copy_hook": "...", "kpi": "..."}},
      {{"data": "18.05", "platforma": "Instagram", "tip": "Engagement", "topic": "...", "format": "Story interactiv", "copy_hook": "...", "kpi": "..."}}
    ]
  }}
}}
"""

    text = gemini_generate(prompt)

    # Curata markdown
    if "```" in text:
        parts = text.split("```")
        for part in parts:
            if part.startswith("json"):
                text = part[4:]
                break
            elif "{" in part:
                text = part
                break
    text = text.strip()

    # Curata caractere de control invalide din JSON (newlines in string-uri)
    import json, re
    # Inlocuieste newlines/tabs neescapate din interiorul string-urilor JSON
    def fix_json_string(s):
        result = []
        in_string = False
        i = 0
        while i < len(s):
            c = s[i]
            if c == '"' and (i == 0 or s[i-1] != '\\'):
                in_string = not in_string
                result.append(c)
            elif in_string and c == '\n':
                result.append('\\n')
            elif in_string and c == '\r':
                result.append('\\r')
            elif in_string and c == '\t':
                result.append('\\t')
            elif in_string and ord(c) < 32:
                result.append(' ')
            else:
                result.append(c)
            i += 1
        return ''.join(result)

    try:
        return json.loads(text)
    except json.JSONDecodeError:
        text_fixed = fix_json_string(text)
        return json.loads(text_fixed)


# ─── TELEGRAM HANDLERS ─────────────────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "👋 Salut! Sunt botul pentru generarea proiectului de examen la *Managementul Vizibilității*.\n\n"
        "Voi genera un Excel complet pe baza organizației tale, gata de predat.\n\n"
        "Câteva întrebări rapide și în ~2 minute primești fișierul.\n\n"
        "Cum te numești? *(Numele de familie)*",
        parse_mode="Markdown"
    )
    return ASK_NAME


async def ask_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["last_name"] = update.message.text.strip()
    await update.message.reply_text("✍️ Prenumele?")
    return ASK_FIRSTNAME


async def ask_firstname(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["first_name"] = update.message.text.strip()
    await update.message.reply_text("🏢 Numele organizației pe care o analizezi?")
    return ASK_ORGANIZATION


async def ask_organization(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["org_name"] = update.message.text.strip()
    await update.message.reply_text(
        "📂 Domeniul de activitate al organizației?\n"
        "_(Ex: IT, Educație, Sănătate, Agricultură, ONG, Mediu etc.)_",
        parse_mode="Markdown"
    )
    return ASK_DOMAIN


async def ask_domain(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["domain"] = update.message.text.strip()
    keyboard = [
        [InlineKeyboardButton("ONG / Asociație", callback_data="ONG")],
        [InlineKeyboardButton("Companie privată / SRL", callback_data="Companie privată (SRL)")],
        [InlineKeyboardButton("Start-up", callback_data="Start-up")],
        [InlineKeyboardButton("Instituție publică", callback_data="Instituție publică")],
        [InlineKeyboardButton("Social Enterprise", callback_data="Social Enterprise")],
    ]
    await update.message.reply_text(
        "🏗️ Tipul organizației:",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ASK_ORG_TYPE


async def ask_org_type(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["org_type"] = query.data
    await query.message.reply_text(
        "🌍 Descrie pe scurt piața / contextul organizației.\n"
        "_(Ex: piața educațională din Moldova, sectorul IT din Chișinău, ONG-uri de mediu din regiune etc.)_",
        parse_mode="Markdown"
    )
    return ASK_MARKET


async def ask_market(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["market"] = update.message.text.strip()
    await update.message.reply_text(
        "🏆 Care sunt 2-3 organizații concurente?\n"
        "_(Scrie numele lor, sau tastează 'nu știu' dacă nu ai)_"
    )
    return ASK_COMPETITORS


async def ask_competitors(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["competitors"] = update.message.text.strip()
    keyboard = [
        [InlineKeyboardButton("Facebook + Instagram", callback_data="Facebook, Instagram")],
        [InlineKeyboardButton("Facebook + Instagram + LinkedIn", callback_data="Facebook, Instagram, LinkedIn")],
        [InlineKeyboardButton("Toate (FB + IG + LI + TikTok + YouTube)", callback_data="Facebook, Instagram, LinkedIn, TikTok, YouTube")],
        [InlineKeyboardButton("Doar website + LinkedIn", callback_data="Website, LinkedIn")],
    ]
    await update.message.reply_text(
        "📱 Pe ce platforme este activă organizația?",
        reply_markup=InlineKeyboardMarkup(keyboard)
    )
    return ASK_PLATFORMS


async def ask_platforms(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    context.user_data["platforms"] = query.data
    await query.message.reply_text("🎓 Grupa ta? _(Ex: MPA-241, MP-23 etc.)_", parse_mode="Markdown")
    return ASK_GRUPA


async def ask_grupa(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["grupa"] = update.message.text.strip()

    sd = context.user_data
    summary = (
        f"✅ *Verifică datele înainte de generare:*\n\n"
        f"👤 Nume: {sd['last_name']} {sd['first_name']}\n"
        f"🎓 Grupa: {sd['grupa']}\n"
        f"🏢 Organizație: {sd['org_name']}\n"
        f"📂 Domeniu: {sd['domain']}\n"
        f"🏗️ Tip: {sd['org_type']}\n"
        f"🌍 Piața: {sd['market']}\n"
        f"🏆 Concurenți: {sd['competitors']}\n"
        f"📱 Platforme: {sd['platforms']}\n\n"
        f"Totul e corect?"
    )
    keyboard = [
        [InlineKeyboardButton("✅ Da, generează!", callback_data="confirm_yes")],
        [InlineKeyboardButton("❌ Nu, reîncepe", callback_data="confirm_no")],
    ]
    await update.message.reply_text(summary, parse_mode="Markdown", reply_markup=InlineKeyboardMarkup(keyboard))
    return CONFIRM


async def confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    if query.data == "confirm_no":
        await query.message.reply_text("Ok, reîncepe cu /start")
        return ConversationHandler.END

    await query.message.reply_text(
        "⏳ *Generez proiectul tău...*\n\n"
        "🔄 Gemini construiește conținut pentru toate cele 11 sheet-uri.\n"
        "Durează ~60-90 secunde. Te rog să aștepți.",
        parse_mode="Markdown"
    )

    student_data = context.user_data.copy()
    user_id = query.from_user.id

    try:
        # Generare cu Gemini
        generated = generate_with_gemini(student_data)

        # Export Excel
        output_path = OUTPUT_DIR / f"examen_{student_data['last_name']}_{student_data['org_name'].replace(' ', '_')}_{user_id}.xlsx"
        fill_excel(student_data, generated, output_path)

        # Trimite fișierul
        with open(output_path, "rb") as f:
            await query.message.reply_document(
                document=f,
                filename=f"Proiect_Vizibilitate_{student_data['last_name']}_{student_data['org_name'][:20]}.xlsx",
                caption=(
                    f"✅ *Proiectul tău este gata!*\n\n"
                    f"📋 Toate cele 11 sheet-uri au fost completate.\n\n"
                    f"⚠️ *Ce mai trebuie să faci manual:*\n"
                    f"• Inserează screenshot-uri reale din paginile sociale (Excel → Insert → Pictures)\n"
                    f"• Completează linkurile reale ale postărilor performante\n"
                    f"• Creează cele 4 materiale în Canva și inserează-le\n"
                    f"• Verifică datele studentului din sheet-ul GHID\n\n"
                    f"📌 Recitește proiectul înainte de predare — personalizează unde simți că e prea generic!"
                ),
                parse_mode="Markdown"
            )

        # Curăță fișierul după trimitere
        output_path.unlink(missing_ok=True)

    except Exception as e:
        logger.error(f"Eroare generare: {e}", exc_info=True)
        context.user_data["_saved"] = student_data
        keyboard = [[InlineKeyboardButton("🔄 Încearcă din nou", callback_data="retry_generate")]]
        await query.message.reply_text(
            f"❌ Eroare la generare: `{str(e)[:200]}`\n\nApasă butonul de mai jos pentru a reîncerca fără a completa din nou datele.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )

    return ConversationHandler.END


async def retry_generate(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    student_data = context.user_data.get("_saved")
    if not student_data:
        await query.message.reply_text("❌ Sesiunea a expirat. Folosește /start pentru a reîncepe.")
        return
    await query.message.reply_text("🔄 Reîncerc generarea... Te rog să aștepți ~60-90 sec.")
    user_id = query.from_user.id
    try:
        generated = generate_with_gemini(student_data)
        output_path = OUTPUT_DIR / f"examen_{student_data['last_name']}_{student_data['org_name'].replace(' ', '_')}_{user_id}.xlsx"
        fill_excel(student_data, generated, output_path)
        with open(output_path, "rb") as f:
            await query.message.reply_document(
                document=f,
                filename=f"Proiect_Vizibilitate_{student_data['last_name']}_{student_data['org_name'][:20]}.xlsx",
                caption=(
                    f"✅ *Proiectul tău este gata!*\n\n"
                    f"📋 Toate cele 11 sheet-uri au fost completate.\n\n"
                    f"⚠️ *Ce mai trebuie să faci manual:*\n"
                    f"• Inserează screenshot-uri reale\n"
                    f"• Completează linkurile reale\n"
                    f"• Creează cele 4 materiale în Canva\n"
                    f"• Verifică datele din sheet-ul GHID\n\n"
                    f"📌 Recitește proiectul înainte de predare!"
                ),
                parse_mode="Markdown"
            )
        output_path.unlink(missing_ok=True)
        context.user_data.pop("_saved", None)
    except Exception as e:
        logger.error(f"Eroare retry: {e}", exc_info=True)
        keyboard = [[InlineKeyboardButton("🔄 Încearcă din nou", callback_data="retry_generate")]]
        await query.message.reply_text(
            f"❌ Eroare din nou: `{str(e)[:200]}`\n\nPoți reîncerca oricând.",
            parse_mode="Markdown",
            reply_markup=InlineKeyboardMarkup(keyboard)
        )


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Anulat. Folosește /start pentru a reîncepe.")
    return ConversationHandler.END


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🤖 *Bot Examen Managementul Vizibilității*\n\n"
        "Comenzi:\n"
        "/start — începe generarea proiectului\n"
        "/cancel — anulează procesul curent\n"
        "/help — acest mesaj\n\n"
        "📧 Probleme? Contactează administratorul grupei.",
        parse_mode="Markdown"
    )


# ─── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    app = Application.builder().token(TELEGRAM_TOKEN).build()

    conv_handler = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            ASK_NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_name)],
            ASK_FIRSTNAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_firstname)],
            ASK_ORGANIZATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_organization)],
            ASK_DOMAIN: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_domain)],
            ASK_ORG_TYPE: [CallbackQueryHandler(ask_org_type)],
            ASK_MARKET: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_market)],
            ASK_COMPETITORS: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_competitors)],
            ASK_PLATFORMS: [CallbackQueryHandler(ask_platforms)],
            ASK_GRUPA: [MessageHandler(filters.TEXT & ~filters.COMMAND, ask_grupa)],
            CONFIRM: [CallbackQueryHandler(confirm)],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv_handler)
    app.add_handler(CallbackQueryHandler(retry_generate, pattern="^retry_generate$"))
    app.add_handler(CommandHandler("help", help_cmd))

    logger.info("Bot pornit!")
    app.run_polling(allowed_updates=Update.ALL_TYPES, drop_pending_updates=True)


if __name__ == "__main__":
    main()
